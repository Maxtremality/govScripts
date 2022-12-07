from win32com.client import Dispatch
import openpyxl
import asyncio

import settings
from ets_data import car_actual_get
from ets_data import token_get


async def sheet_copy(sheet_name: str, file_path: str):
    xl = Dispatch("Excel.Application")
    xl.ScreenUpdating = False
    xl.DisplayAlerts = False
    xl.EnableEvents = False
    xl.Visible = False

    xl_sample = xl.Workbooks.Open('C:/Users/maxtremality/Documents/govScripts/new.xlsx')
    xl_final = xl.Workbooks.Open(file_path)
    ws = xl_sample.Worksheets(sheet_name)
    ws.Copy(Before=xl_final.Worksheets(1))

    xl_final.Close(SaveChanges=True)
    xl.Quit()


async def main():
    token = await token_get(settings.TRAVEL_TIME_ACCESS['avdSAO']['login'], settings.TRAVEL_TIME_ACCESS['avdSAO']['password'])
    car_actual = await car_actual_get(token, '')

    sample = openpyxl.load_workbook('new.xlsx')

    print(sample['КМ']['A5'].value)

    for x in car_actual['result']['rows']:
        if x['company_structure_name'] == 'ДЭУ-1':
            if x['odometer_mileage'] != None and x['motohours_equip_mileage'] != None:
                await sheet_copy('СпецОб', 'C:/Users/maxtremality/Documents/govScripts/ДЭУ-1.xlsx')
                wb = openpyxl.load_workbook('ДЭУ-1.xlsx')
                wb['СпецОб'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ДЭУ-1.xlsx')
            elif x['odometer_mileage'] != None and x['motohours_equip_mileage'] == None:
                await sheet_copy('КМ', 'C:/Users/maxtremality/Documents/govScripts/ДЭУ-1.xlsx')
                wb = openpyxl.load_workbook('ДЭУ-1.xlsx')
                wb['КМ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ДЭУ-1.xlsx')
            elif x['odometer_mileage'] == None and x['motohours_mileage'] != None:
                await sheet_copy('МЧ', 'C:/Users/maxtremality/Documents/govScripts/ДЭУ-1.xlsx')
                wb = openpyxl.load_workbook('ДЭУ-1.xlsx')
                wb['МЧ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ДЭУ-1.xlsx')
        if x['company_structure_name'] == 'ДЭУ-2':
            if x['odometer_mileage'] != None and x['motohours_equip_mileage'] != None:
                await sheet_copy('СпецОб', 'C:/Users/maxtremality/Documents/govScripts/ДЭУ-2.xlsx')
                wb = openpyxl.load_workbook('ДЭУ-2.xlsx')
                wb['СпецОб'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ДЭУ-2.xlsx')
            elif x['odometer_mileage'] != None and x['motohours_equip_mileage'] == None:
                await sheet_copy('КМ', 'C:/Users/maxtremality/Documents/govScripts/ДЭУ-2.xlsx')
                wb = openpyxl.load_workbook('ДЭУ-2.xlsx')
                wb['КМ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ДЭУ-2.xlsx')
            elif x['odometer_mileage'] == None and x['motohours_mileage'] != None:
                await sheet_copy('МЧ', 'C:/Users/maxtremality/Documents/govScripts/ДЭУ-2.xlsx')
                wb = openpyxl.load_workbook('ДЭУ-2.xlsx')
                wb['МЧ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ДЭУ-2.xlsx')
        if x['company_structure_name'] == 'ДЭУ-3':
            if x['odometer_mileage'] != None and x['motohours_equip_mileage'] != None:
                await sheet_copy('СпецОб', 'C:/Users/maxtremality/Documents/govScripts/ДЭУ-3.xlsx')
                wb = openpyxl.load_workbook('ДЭУ-3.xlsx')
                wb['СпецОб'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ДЭУ-3.xlsx')
            elif x['odometer_mileage'] != None and x['motohours_equip_mileage'] == None:
                await sheet_copy('КМ', 'C:/Users/maxtremality/Documents/govScripts/ДЭУ-3.xlsx')
                wb = openpyxl.load_workbook('ДЭУ-3.xlsx')
                wb['КМ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ДЭУ-3.xlsx')
            elif x['odometer_mileage'] == None and x['motohours_mileage'] != None:
                await sheet_copy('МЧ', 'C:/Users/maxtremality/Documents/govScripts/ДЭУ-3.xlsx')
                wb = openpyxl.load_workbook('ДЭУ-3.xlsx')
                wb['МЧ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ДЭУ-3.xlsx')
        if x['company_structure_name'] == 'БРТС':
            if x['odometer_mileage'] != None and x['motohours_equip_mileage'] != None:
                await sheet_copy('СпецОб', 'C:/Users/maxtremality/Documents/govScripts/БРТС.xlsx')
                wb = openpyxl.load_workbook('БРТС.xlsx')
                wb['СпецОб'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('БРТС.xlsx')
            elif x['odometer_mileage'] != None and x['motohours_equip_mileage'] == None:
                await sheet_copy('КМ', 'C:/Users/maxtremality/Documents/govScripts/БРТС.xlsx')
                wb = openpyxl.load_workbook('БРТС.xlsx')
                wb['КМ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('БРТС.xlsx')
            elif x['odometer_mileage'] == None and x['motohours_mileage'] != None:
                await sheet_copy('МЧ', 'C:/Users/maxtremality/Documents/govScripts/БРТС.xlsx')
                wb = openpyxl.load_workbook('БРТС.xlsx')
                wb['МЧ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('БРТС.xlsx')
        if x['company_structure_name'] == 'Спецавтоколонна':
            if x['odometer_mileage'] != None and x['motohours_equip_mileage'] != None:
                await sheet_copy('СпецОб', 'C:/Users/maxtremality/Documents/govScripts/Спецавтоколонна.xlsx')
                wb = openpyxl.load_workbook('Спецавтоколонна.xlsx')
                wb['СпецОб'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('Спецавтоколонна.xlsx')
            elif x['odometer_mileage'] != None and x['motohours_equip_mileage'] == None:
                await sheet_copy('КМ', 'C:/Users/maxtremality/Documents/govScripts/Спецавтоколонна.xlsx')
                wb = openpyxl.load_workbook('Спецавтоколонна.xlsx')
                wb['КМ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('Спецавтоколонна.xlsx')
            elif x['odometer_mileage'] == None and x['motohours_mileage'] != None:
                await sheet_copy('МЧ', 'C:/Users/maxtremality/Documents/govScripts/Спецавтоколонна.xlsx')
                wb = openpyxl.load_workbook('Спецавтоколонна.xlsx')
                wb['МЧ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('Спецавтоколонна.xlsx')
        if x['company_structure_name'] == 'ОАСГ':
            if x['odometer_mileage'] != None and x['motohours_equip_mileage'] != None:
                await sheet_copy('СпецОб', 'C:/Users/maxtremality/Documents/govScripts/ОАСГ.xlsx')
                wb = openpyxl.load_workbook('ОАСГ.xlsx')
                wb['СпецОб'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ОАСГ.xlsx')
            elif x['odometer_mileage'] != None and x['motohours_equip_mileage'] == None:
                await sheet_copy('КМ', 'C:/Users/maxtremality/Documents/govScripts/ОАСГ.xlsx')
                wb = openpyxl.load_workbook('ОАСГ.xlsx')
                wb['КМ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ОАСГ.xlsx')
            elif x['odometer_mileage'] == None and x['motohours_mileage'] != None:
                await sheet_copy('МЧ', 'C:/Users/maxtremality/Documents/govScripts/ОАСГ.xlsx')
                wb = openpyxl.load_workbook('ОАСГ.xlsx')
                wb['МЧ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ОАСГ.xlsx')
        if x['company_structure_name'] == 'СНОС':
            if x['odometer_mileage'] != None and x['motohours_equip_mileage'] != None:
                await sheet_copy('СпецОб', 'C:/Users/maxtremality/Documents/govScripts/СНОС.xlsx')
                wb = openpyxl.load_workbook('СНОС.xlsx')
                wb['СпецОб'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('СНОС.xlsx')
            elif x['odometer_mileage'] != None and x['motohours_equip_mileage'] == None:
                await sheet_copy('КМ', 'C:/Users/maxtremality/Documents/govScripts/СНОС.xlsx')
                wb = openpyxl.load_workbook('СНОС.xlsx')
                wb['КМ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('СНОС.xlsx')
            elif x['odometer_mileage'] == None and x['motohours_mileage'] != None:
                await sheet_copy('МЧ', 'C:/Users/maxtremality/Documents/govScripts/СНОС.xlsx')
                wb = openpyxl.load_workbook('СНОС.xlsx')
                wb['МЧ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('СНОС.xlsx')
        if x['company_structure_name'] == 'ДРУ':
            if x['odometer_mileage'] != None and x['motohours_equip_mileage'] != None:
                await sheet_copy('СпецОб', 'C:/Users/maxtremality/Documents/govScripts/ДРУ.xlsx')
                wb = openpyxl.load_workbook('ДРУ.xlsx')
                wb['СпецОб'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ДРУ.xlsx')
            elif x['odometer_mileage'] != None and x['motohours_equip_mileage'] == None:
                await sheet_copy('КМ', 'C:/Users/maxtremality/Documents/govScripts/ДРУ.xlsx')
                wb = openpyxl.load_workbook('ДРУ.xlsx')
                wb['КМ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ДРУ.xlsx')
            elif x['odometer_mileage'] == None and x['motohours_mileage'] != None:
                await sheet_copy('МЧ', 'C:/Users/maxtremality/Documents/govScripts/ДРУ.xlsx')
                wb = openpyxl.load_workbook('ДРУ.xlsx')
                wb['МЧ'].title = x['gov_number']
                wb[x['gov_number']]['A5'] = 'Карточка учета ' + x['gov_number']
                wb.save('ДРУ.xlsx')

if __name__ == "__main__":
    asyncio.run(main())