import asyncio
from openpyxl import load_workbook

import settings
from ets_data import token_get
from ets_data import fuel_consumption_rate_post
from ets_data import fuel_consumption_rate_get
from ets_data import fuel_consumption_rate_delete
from ets_data import car_actual_get


async def main():
    token = await token_get(settings.TRAVEL_TIME_ACCESS['avdSAO']['login'], settings.TRAVEL_TIME_ACCESS['avdSAO']['password'])

    car_actual = await car_actual_get(token, '')
    file = open('fuel_norm_base64.txt')
    fuel_consumption_rate = await fuel_consumption_rate_get(token)

    wb = load_workbook('Нормы топлива 2022 - 2023 500.xlsx')
    for x in range(wb['data'].max_row):
        for car in car_actual['result']['rows']:
            if car['gov_number'] == wb['data']['A' + str(x+1)].value:
                if wb['data']['E' + str(x+1)].value == 'Холостой ход' and wb['data']['F' + str(x+1)].value == 'л/км':
                    operation_id = 598
                elif wb['data']['E' + str(x+1)].value == 'Рабочий ход' and wb['data']['F' + str(x+1)].value == 'л/км':
                    operation_id = 594
                elif wb['data']['E' + str(x+1)].value == 'Холостой ход' and wb['data']['F' + str(x+1)].value == 'л/моточас':
                    operation_id = 1238
                elif wb['data']['E' + str(x+1)].value == 'Рабочий ход' and wb['data']['F' + str(x+1)].value == 'л/моточас':
                    operation_id = 1224
                elif wb['data']['E' + str(x+1)].value == 'Работа на месте' and wb['data']['F' + str(x+1)].value == 'л/час':
                    operation_id = 1234
                elif wb['data']['E' + str(x+1)].value == 'Автономный двигатель' and wb['data']['F' + str(x+1)].value == 'л/моточас' and wb['data']['G' + str(x+1)].value == 'true':
                    operation_id = 1213
                elif wb['data']['E' + str(x+1)].value == 'Автономный двигатель' and wb['data']['F' + str(x+1)].value == 'л/моточас' and wb['data']['G' + str(x+1)].value == 'false':
                    operation_id = 1215

                data = {
                    'summer_rate': wb['data']['H' + str(x+1)].value,
                    'winter_rate': wb['data']['I' + str(x+1)].value,
                    'car_id': car['asuods_id'],
                    'car_model_id': car['model_id'],
                    'car_special_model_id': car['special_model_id'],
                    'comment': wb['data']['J' + str(x+1)].value,
                    'operation_id': operation_id,
                    'order_date': '2022-10-28',
                    'order_number': '184 о/д',
                    'engine_kind_id': 1,
                    'files': [{
                        'action': 'add',
                        'kind': 'order',
                        'name': 'Нормы топлива от 28.10.2022.pdf',
                        'nativeFile': {},
                        'base64': file.read()
                    }],
                    'full_model_name': None,
                    'company_id': None,
                    'company_name': None,
                    'body_capacity': None,
                    'company_structure_id': None,
                    'company_structure_name': None,
                    'engine_kind_name': None,
                    'gov_number': None,
                    'id': None,
                    'is_excluding_mileage': None,
                    'load_capacity': None,
                    'max_speed': None,
                    'measure_unit_id': None,
                    'measure_unit_name': None,
                    'model_name': None,
                    'okrug_name': None,
                    'operation_equipment': None,
                    'operation_name': None
                }

                await fuel_consumption_rate_post(token, data)

                # for norm in fuel_consumption_rate['result']['rows']:
                #     if norm['gov_number'] == 'У817РР799':
                #         await fuel_consumption_rate_delete(token, norm['id'])

if __name__ == "__main__":
    asyncio.run(main())
