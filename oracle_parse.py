import openpyxl
import io
import time
import asyncio
import aiohttp
import requests
from datetime import datetime, timedelta

import settings
from db2 import data_get
from db2 import data_post
from db2 import data_truncate
from db2 import data_delete


async def oracle_login(username, password):
    url = 'http://185.173.2.48/analytics/saw.dll?BIEEHome'
    payload = {
        'NQUser': username,
        'NQPassword': password,
        'Locale': None,
        'startPage': 1,
        'icharset': 'utf-8'
    }
    response = requests.get(url)
    LBINFO = response.headers['Set-Cookie'].split(";")[0]
    headers = {
        'Cookie': response.headers['Set-Cookie'].split(";")[0],
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    }
    response = requests.post(url, data=payload, headers=headers, allow_redirects=False)
    NQID = response.headers['Set-Cookie'].split(";")[0]
    return {
        'Cookie': f'{ LBINFO };{ NQID }',
        'Content-Type': 'application/x-www-form-urlencoded'
    }


async def get_oracle_xlsx(session, report_name, headers, payload):
    url = 'http://185.173.2.48/xmlpserver/servlet/xdo'

    async with session.post(url, headers=headers, data=payload, timeout=600) as response:
        if response.status == 200:
            response = await response.content.read()
            print(f'{report_name} загружен')
            return {'name': report_name, 'data': response}
        else:
            print("Ошибка при выполнении запроса:", response.status)


async def main():
    start = time.time()

    headers = await oracle_login('ASBURMISTROV', 'ASBURMISTROV123')
    interval = await data_get('from_dt, to_dt', 'api_cartraveltimeform', '', settings.DB_API, 'public', settings.DB_ACCESS)
    payload_data = {
        'Статистика перемещения техники':
            {
                '_xpf': 1,
                '_xpt': 0,
                '_xdo': '/Publisher/Аналитическая группа/Телеметрия/Отчет «Статистика перемещения техники».xdo',
                '_xmode': 2,
                'xdo:xdo:_paramsp_source_div_input': 'ОДХ',
                '_paramsp_source': 'UDO',
                'xdo:xdo:_paramsp_region_id_tab_div_input': 'ВАО',
                '_paramsp_region_id_tab': 11,
                'xdo:xdo:_paramsp_date_param_div_input': 'Уборочные сутки',
                '_paramsp_date_param': 'cleaning_date',
                '_paramsp_date_from': interval[-1][0].strftime('%d-%m-%Y'),
                '_paramsp_date_to': interval[-1][1].strftime('%d-%m-%Y'),
                '_paramsp_time_from': interval[-1][0].strftime('%H:%M'),
                '_paramsp_time_to': interval[-1][1].strftime('%H:%M'),
                'xdo:xdo:_paramsp_customer_id_tab_div_input': 'АвД ВАО',
                '_paramsp_customer_id_tab': 9000019,
                'xdo:xdo:_paramsp_contractor_id_div_input': 'Все',
                '_paramsp_contractor_id': '*',
                'xdo:xdo:_paramsp_owner_id_tab_div_input': 'Все',
                '_paramsp_owner_id_tab': '*',
                'xdo:xdo:_paramsp_func_type_group_id_div_input': 'Уборочная техника',
                '_paramsp_func_type_group_id': 1,
                'xdo:xdo:_paramsp_func_type_id_div_input': 'Все',
                '_paramsp_func_type_id': '*',
                '_xt': 'отчет',
                '_xf': 'xlsx',
                '_xautorun': 'false'
            },
        'Посещение объектов ПЧ':
            {
                '_xpf': None,
                '_xpt': 0,
                '_xdo': '/Publisher/Монитор/Посещение объектов/Отчет Посещение объектов уборочной техникой по заказчику учредителю.xdo',
                '_xmode': 2,
                '_paramsp_target_date': (datetime.now() - timedelta(hours=9)).strftime('%d-%m-%Y'),
                'xdo:xdo:_paramsp_subject_contract_div_input': 'ОДХ',
                '_paramsp_subject_contract': 1,
                'xdo:xdo:_paramsp_element_div_input': 'Проезжая часть',
                '_paramsp_element': 1,
                'xdo:xdo:_paramsp_customer_id_div_input': 'АвД ВАО',
                '_paramsp_customer_id': 9000019,
                'xdo:xdo:_paramsp_okrug_id_div_input': 'ВАО',
                '_paramsp_okrug_id': 11,
                '_xt': 'отчет',
                '_xf': 'xlsx',
                '_xautorun': 'false',
            },
        'Посещение объектов Тротуары':
            {
                '_xpf': None,
                '_xpt': 0,
                '_xdo': '/Publisher/Монитор/Посещение объектов/Отчет Посещение объектов уборочной техникой по заказчику учредителю.xdo',
                '_xmode': 2,
                '_paramsp_target_date': (datetime.now() - timedelta(hours=9)).strftime('%d-%m-%Y'),
                'xdo:xdo:_paramsp_subject_contract_div_input': 'ОДХ',
                '_paramsp_subject_contract': 1,
                'xdo:xdo:_paramsp_element_div_input': 'Тротуар',
                '_paramsp_element': 2,
                'xdo:xdo:_paramsp_customer_id_div_input': 'АвД ВАО',
                '_paramsp_customer_id': 9000019,
                'xdo:xdo:_paramsp_okrug_id_div_input': 'ВАО',
                '_paramsp_okrug_id': 11,
                '_xt': 'отчет',
                '_xf': 'xlsx',
                '_xautorun': 'false',
            },
    }

    async with aiohttp.ClientSession() as session:
        tasks = []
        for report_name in payload_data:
            tasks.append(asyncio.ensure_future(get_oracle_xlsx(session, report_name, headers, payload_data[report_name])))

        files = await asyncio.gather(*tasks)

        for file in files:
            data = []
            wb = openpyxl.load_workbook(filename=io.BytesIO(file['data']))
            sheet = wb.active

            for _ in range(5):
                sheet.delete_rows(1)
            sheet.delete_rows(sheet.max_row)

            if file['name'] == 'Посещение объектов ПЧ' or file['name'] == 'Посещение объектов Тротуары':
                sheet.delete_cols(32, 33)
                for row in sheet.iter_rows():
                    data.append(((datetime.now()).isoformat(), (datetime.now() - timedelta(hours=9)).strftime('%Y-%m-%d')) + tuple(0 if cell.value is None else cell.value for cell in row))
                columns = 'datetime_upload, date, object_id, name, okrug, customer_id, customer, contractor_id, contractor, category,' \
                          'subcategory, area, fact_traveled_exec, norm_traveled_exec, traveled_area, plan_area, procent, area_distance,' \
                          'area_gutter, fact_distance_gutter_exec, fact_distance_exec, fact_gutter_exec, norm_distance_exec, norm_gutter_exec,' \
                          'plan_distance_gutter_exec, plan_distance_exec, plan_gutter_exec, traveled_area_distance_gutter, traveled_area_distance,' \
                          'traveled_area_gutter, procent_distance_gutter, procent_all, status'
                params = f"where extract(day from ('{ (datetime.now() - timedelta(hours=9)).date() }'::timestamp without time zone - date::timestamp without time zone)) <= 0"
                if file['name'] == 'Посещение объектов ПЧ' and data:
                    await data_delete('oracle_monitor_roadway', params, settings.DB_SCRIPTS, 'scripts', settings.DB_ACCESS)
                    await data_post(data, 'oracle_monitor_roadway', columns, settings.DB_SCRIPTS, 'scripts', settings.DB_ACCESS)
                elif file['name'] == 'Посещение объектов Тротуары' and data:
                    await data_delete('oracle_monitor_footway', params, settings.DB_SCRIPTS, 'scripts', settings.DB_ACCESS)
                    await data_post(data, 'oracle_monitor_footway', columns, settings.DB_SCRIPTS, 'scripts', settings.DB_ACCESS)
            elif file['name'] == 'Статистика перемещения техники':
                sheet.delete_cols(18, 32)
                for row in sheet.iter_rows():
                    data.append((interval[-1][0].isoformat(), interval[-1][1].isoformat()) + tuple(0 if cell.value is None else cell.value for cell in row))
                columns = 'from_dt, to_dt, customer, contractor, car_type, gov_number, bnso_owner, bnso_provider, category, object_id, object_name, ' \
                          'arrival, departure, travel_time, avg_speed, max_speed, distance, car_group, car_owner'
                if data:
                    await data_truncate('oracle_movement_statistics', settings.DB_SCRIPTS, 'scripts', settings.DB_ACCESS)
                    await data_post(data, 'oracle_movement_statistics', columns, settings.DB_SCRIPTS, 'scripts', settings.DB_ACCESS)

    print('Время выполнения:', time.strftime("%H:%M:%S", time.gmtime(time.time() - start)))

if __name__ == '__main__':
    asyncio.run(main())
