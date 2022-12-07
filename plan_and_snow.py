from openpyxl import load_workbook
import shutil
from glob import glob


def main():
    # shutil.copy2('техника+снег САО шаблон.xlsx', 'техника+снег САО ' + 'готовый' + '.xlsx')

    result_wb = load_workbook('техника+снег САО готовый.xlsx')
    result_ws = result_wb[result_wb.sheetnames[0]]
    f = 1

    for file in glob('*.xlsx'):
        if file == 'техника+снег САО готовый.xlsx' or file == 'техника+снег САО шаблон.xlsx' or '~$' in file:
            continue
        wb = load_workbook(file)
        ws = wb[wb.sheetnames[0]]
        print(file)

        if 'АвД САО' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D11'] = ws['D11'].value
            result_ws['I11'] = ws['H11'].value
            result_ws['N11'] = ws['L11'].value
            result_ws['S11'] = ws['P11'].value
            result_ws['X11'] = ws['T11'].value
            result_ws['AC11'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Аэропорт' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D12'] = ws['D11'].value
            result_ws['I12'] = ws['H11'].value
            result_ws['N12'] = ws['L11'].value
            result_ws['S12'] = ws['P11'].value
            result_ws['X12'] = ws['T11'].value
            result_ws['AC12'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Беговой' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D13'] = ws['D11'].value
            result_ws['I13'] = ws['H11'].value
            result_ws['N13'] = ws['L11'].value
            result_ws['S13'] = ws['P11'].value
            result_ws['X13'] = ws['T11'].value
            result_ws['AC13'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Бескудниковский' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D14'] = ws['D11'].value
            result_ws['I14'] = ws['H11'].value
            result_ws['N14'] = ws['L11'].value
            result_ws['S14'] = ws['P11'].value
            result_ws['X14'] = ws['T11'].value
            result_ws['AC14'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Войковский' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D15'] = ws['D11'].value
            result_ws['I15'] = ws['H11'].value
            result_ws['N15'] = ws['L11'].value
            result_ws['S15'] = ws['P11'].value
            result_ws['X15'] = ws['T11'].value
            result_ws['AC15'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Восточное Дегунино' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D16'] = ws['D11'].value
            result_ws['I16'] = ws['H11'].value
            result_ws['N16'] = ws['L11'].value
            result_ws['S16'] = ws['P11'].value
            result_ws['X16'] = ws['T11'].value
            result_ws['AC16'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Головинский' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D17'] = ws['D11'].value
            result_ws['I17'] = ws['H11'].value
            result_ws['N17'] = ws['L11'].value
            result_ws['S17'] = ws['P11'].value
            result_ws['X17'] = ws['T11'].value
            result_ws['AC17'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Дмитровский' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D18'] = ws['D11'].value
            result_ws['I18'] = ws['H11'].value
            result_ws['N18'] = ws['L11'].value
            result_ws['S18'] = ws['P11'].value
            result_ws['X18'] = ws['T11'].value
            result_ws['AC18'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Западное Дегунино' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D19'] = ws['D11'].value
            result_ws['I19'] = ws['H11'].value
            result_ws['N19'] = ws['L11'].value
            result_ws['S19'] = ws['P11'].value
            result_ws['X19'] = ws['T11'].value
            result_ws['AC19'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Коптево' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D20'] = ws['D11'].value
            result_ws['I20'] = ws['H11'].value
            result_ws['N20'] = ws['L11'].value
            result_ws['S20'] = ws['P11'].value
            result_ws['X20'] = ws['T11'].value
            result_ws['AC20'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Левобережный' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D21'] = ws['D11'].value
            result_ws['I21'] = ws['H11'].value
            result_ws['N21'] = ws['L11'].value
            result_ws['S21'] = ws['P11'].value
            result_ws['X21'] = ws['T11'].value
            result_ws['AC21'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Молжаниновский' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D22'] = ws['D11'].value
            result_ws['I22'] = ws['H11'].value
            result_ws['N22'] = ws['L11'].value
            result_ws['S22'] = ws['P11'].value
            result_ws['X22'] = ws['T11'].value
            result_ws['AC22'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Савеловский' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D23'] = ws['D11'].value
            result_ws['I23'] = ws['H11'].value
            result_ws['N23'] = ws['L11'].value
            result_ws['S23'] = ws['P11'].value
            result_ws['X23'] = ws['T11'].value
            result_ws['AC23'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Сокол' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D24'] = ws['D11'].value
            result_ws['I24'] = ws['H11'].value
            result_ws['N24'] = ws['L11'].value
            result_ws['S24'] = ws['P11'].value
            result_ws['X24'] = ws['T11'].value
            result_ws['AC24'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Тимирязевский' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D25'] = ws['D11'].value
            result_ws['I25'] = ws['H11'].value
            result_ws['N25'] = ws['L11'].value
            result_ws['S25'] = ws['P11'].value
            result_ws['X25'] = ws['T11'].value
            result_ws['AC25'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Ховрино' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D26'] = ws['D11'].value
            result_ws['I26'] = ws['H11'].value
            result_ws['N26'] = ws['L11'].value
            result_ws['S26'] = ws['P11'].value
            result_ws['X26'] = ws['T11'].value
            result_ws['AC26'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        elif 'Хорошевский' in file \
            and 'Планируется к работе' in ws['D10'].value \
            and 'Планируется к работе' in ws['H10'].value \
            and 'Планируется к работе' in ws['L10'].value \
            and 'Планируется к работе' in ws['P10'].value \
            and 'Планируется к работе' in ws['T10'].value \
            and 'Планируется к работе' in ws['X10'].value:
            result_ws['D27'] = ws['D11'].value
            result_ws['I27'] = ws['H11'].value
            result_ws['N27'] = ws['L11'].value
            result_ws['S27'] = ws['P11'].value
            result_ws['X27'] = ws['T11'].value
            result_ws['AC27'] = ws['X11'].value
            print(file, 'записан')
            f += 1

        wb.close()

    result_wb.save('техника+снег САО готовый.xlsx')
    result_wb.close()

    if f < 17:
        print('Загружено', f, 'файлов \n',
              'Загружены не все файлы или при выполнении возникла ошибка. Проверьте соответвие файлов: \n',
              '1. Файлы должны иметь расширение xlsx \n',
              '2. В наименовании должно содержать название организации с учетом регистра (АвД САО, Аэропорт и т.д.) \n',
              '3. Первый лист должен быть с данными \n',
              '4. Очередность наименований организаций должна совпадать с шаблоном \n',
              '5. Адреса ячеек должны совпадать с шаблоном (Адерса это A1, B10, F12 и т.д.) \n \n',
              'Для повторной записи, удалите из папки файлы с пометкой "готовый"')
    else:
        print('Все данные успешно загружены')

    print('Нажмите Enter для закрытия окна')
    input()


if __name__ == "__main__":
    main()
